"""
PVD AA Cabin Report PDF Generator — Cloud Version
Reads Excel from SharePoint via Microsoft Graph API.
Generates two PDFs and saves them to SharePoint.
Sends email via Microsoft Graph API.
"""

import os
import io
import sys
import json
import base64
import requests
import openpyxl
from datetime import datetime, date, time
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether
)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER

# ─── MICROSOFT GRAPH AUTH CONFIG (set via GitHub Actions secrets) ─────────────
TENANT_ID     = os.environ["TENANT_ID"]
CLIENT_ID     = os.environ["CLIENT_ID"]
CLIENT_SECRET = os.environ["CLIENT_SECRET"]

# ─── SHAREPOINT CONFIG ────────────────────────────────────────────────────────
# Drive ID for foxtrotaviationcom.sharepoint.com/sites/DataHub
DRIVE_ID = "b!_bzXaIx86kOufgJN3ih-BaDIDthKYuxJkJtLi1Bm5irGjCEnK-VHSpBRRm3_SDKU"

# Path to the PVD Excel file (relative to drive root)
EXCEL_PATH = "/Report Automation/PVD AA Cabin/PVD Tables.xlsx"

# SharePoint folder path where PDFs will be saved
PDF_OUTPUT_FOLDER = "/Report Automation/PVD AA Cabin/Daily Reports"

# SharePoint path to the screenshot embedded in the email body
SCREENSHOT_PATH = "/Flow Dumps/AA Screenshots/PVD.png"

# ─── EMAIL CONFIG ─────────────────────────────────────────────────────────────
# Set to False to skip sending the email (useful when testing)
SEND_EMAIL = True

EMAIL_FROM = "foxtrot.automation@foxtrotaviation.com"

EMAIL_RECIPIENTS = [
    "jamie.mathers@foxtrotaviation.com",
    "joya.gregory@foxtrotaviation.com",
    "ricardo.garcia@foxtrotaviation.com",
    "levorn.smallwood@foxtrotaviation.com",
    "lori.desjardin@foxtrotaviation.com",
    "maren.pinpin@foxtrotaviation.com",
    "daniel.digiambattista@foxtrotaviation.com",
    "anthony.pentz@foxtrotaviation.com",
    "chris.stump@foxtrotaviation.com",
    "samuel.kosco@foxtrotaviation.com",
]

# ─── MICROSOFT GRAPH AUTH ─────────────────────────────────────────────────────

def get_access_token():
    """Obtain an OAuth2 access token from Microsoft identity platform."""
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    resp = requests.post(url, data={
        "grant_type":    "client_credentials",
        "client_id":     CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope":         "https://graph.microsoft.com/.default",
    })
    resp.raise_for_status()
    return resp.json()["access_token"]


# ─── SHAREPOINT FILE OPERATIONS ───────────────────────────────────────────────

def download_excel(token):
    """Download the PVD Excel file from SharePoint into a BytesIO buffer."""
    url = (f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}"
           f"/root:{EXCEL_PATH}:/content")
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"})
    resp.raise_for_status()
    print("PVD Excel file downloaded from SharePoint.")
    return io.BytesIO(resp.content)


def download_screenshot(token):
    """Download the station screenshot PNG from SharePoint."""
    url = (f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}"
           f"/root:{SCREENSHOT_PATH}:/content")
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"})
    resp.raise_for_status()
    print("Screenshot downloaded from SharePoint.")
    return resp.content


def upload_pdf(token, filename, pdf_bytes):
    """Upload a PDF to the SharePoint PVD output folder."""
    upload_path = f"{PDF_OUTPUT_FOLDER}/{filename}"
    url = (f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}"
           f"/root:{upload_path}:/content")
    resp = requests.put(
        url,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type":  "application/pdf",
        },
        data=pdf_bytes,
    )
    resp.raise_for_status()
    print(f"Uploaded to SharePoint: {upload_path}")
    return resp.json().get("webUrl", "")


# ─── EMAIL SENDING ────────────────────────────────────────────────────────────

def send_email(token, subject, body_text, recipients, from_address,
               attachments, screenshot_bytes=None):
    """
    Send HTML email via Microsoft Graph API with PDF attachments.
    If screenshot_bytes is provided it is embedded inline in the email body
    below the message text, using a cid: reference.
    attachments: list of (filename, bytes) tuples (PDFs)
    """
    attachment_list = []

    # Inline screenshot — referenced in HTML body via cid:stationScreenshot
    if screenshot_bytes:
        attachment_list.append({
            "@odata.type":  "#microsoft.graph.fileAttachment",
            "name":         "screenshot.png",
            "contentType":  "image/png",
            "contentBytes": base64.b64encode(screenshot_bytes).decode("utf-8"),
            "contentId":    "stationScreenshot",
            "isInline":     True,
        })

    # PDF file attachments (appear as regular attachments)
    for filename, pdf_bytes in attachments:
        attachment_list.append({
            "@odata.type":  "#microsoft.graph.fileAttachment",
            "name":         filename,
            "contentType":  "application/pdf",
            "contentBytes": base64.b64encode(pdf_bytes).decode("utf-8"),
        })

    # Build HTML body: each non-empty line becomes a paragraph, image appended below
    html_paragraphs = "".join(
        f'<p style="margin:0 0 10px 0;">{line}</p>'
        for line in body_text.split("\n") if line.strip()
    )
    img_tag = (
        '<p><img src="cid:stationScreenshot" '
        'style="max-width:600px;width:100%;border:1px solid #ccc;" '
        'alt="Station Screenshot" /></p>'
    ) if screenshot_bytes else ""
    html_body = (
        '<html><body style="font-family:Calibri,Arial,sans-serif;font-size:14px;">'
        f"{html_paragraphs}{img_tag}"
        "</body></html>"
    )

    message = {
        "subject": subject,
        "body":    {"contentType": "HTML", "content": html_body},
        "toRecipients": [
            {"emailAddress": {"address": addr}} for addr in recipients
        ],
        "attachments": attachment_list,
    }

    url = f"https://graph.microsoft.com/v1.0/users/{from_address}/sendMail"
    resp = requests.post(
        url,
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        json={"message": message, "saveToSentItems": True},
    )
    resp.raise_for_status()
    print(f"Email sent to {len(recipients)} recipients.")


# ─── VALUE FORMATTING ─────────────────────────────────────────────────────────

def fmt_value(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%-m/%-d/%Y") if sys.platform != "win32" else val.strftime("%#m/%#d/%Y")
    if isinstance(val, time):
        return val.strftime("%-I:%M %p") if sys.platform != "win32" else val.strftime("%#I:%M %p")
    return str(val)


def fmt_date_short(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%-m/%-d") if sys.platform != "win32" else val.strftime("%#m/%#d")
    return str(val)


# ─── TABLE HELPERS ────────────────────────────────────────────────────────────

def get_table_data(ws, row_start, row_end, col_start, col_end):
    data = []
    for row in ws.iter_rows(
        min_row=row_start, max_row=row_end,
        min_col=col_start, max_col=col_end,
        values_only=True,
    ):
        formatted = [fmt_value(v) for v in row]
        if any(v != "" for v in formatted):
            data.append(formatted)
    return data


def make_table(data, col_widths=None,
               header_bg=colors.HexColor('#2F5496'),
               stripe_color=colors.HexColor('#DCE6F1')):
    if not data:
        return None
    while data and all(v == "" for v in data[-1]):
        data.pop()
    if not data:
        return None

    n_cols     = len(data[0])
    available  = 7.5 * inch
    if col_widths is None:
        col_widths = [available / n_cols] * n_cols

    header_style = ParagraphStyle(
        'TableHeader', fontName='Helvetica-Bold', fontSize=8,
        textColor=colors.white, alignment=TA_CENTER, leading=10,
    )
    cell_style = ParagraphStyle(
        'TableCell', fontName='Helvetica', fontSize=7.5,
        textColor=colors.black, alignment=TA_LEFT, leading=10,
    )

    wrapped = []
    for r_idx, row in enumerate(data):
        new_row = []
        for cell in row:
            if isinstance(cell, str):
                s = header_style if r_idx == 0 else cell_style
                new_row.append(Paragraph(cell, s))
            else:
                new_row.append(cell)
        wrapped.append(new_row)

    t = Table(wrapped, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0),  header_bg),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#B8CCE4')),
        ('BOX',        (0, 0), (-1, -1), 1,   colors.HexColor('#2F5496')),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING',   (0, 0), (-1, -1), 4),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
    ]
    for i in range(1, len(wrapped)):
        if i % 2 == 0:
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), stripe_color))
    t.setStyle(TableStyle(style_cmds))
    return t


# ─── PDF GENERATION ───────────────────────────────────────────────────────────

def build_pdf_bytes(ws_email, ws_jobs, date_line, sections, lms_section,
                    abridged=False):
    """Generate a PDF and return its bytes."""
    buf = io.BytesIO()

    style_title = ParagraphStyle(
        'ReportTitle', fontName='Helvetica-Bold', fontSize=14,
        textColor=colors.HexColor('#1F3864'), spaceAfter=4,
    )
    style_section = ParagraphStyle(
        'SectionHeader', fontName='Helvetica-Bold', fontSize=9,
        textColor=colors.HexColor('#2F5496'), spaceBefore=8, spaceAfter=2,
    )
    style_body = ParagraphStyle(
        'Body', fontName='Helvetica', fontSize=8.5, spaceAfter=4, leading=12,
    )
    style_small = ParagraphStyle(
        'SmallNote', fontName='Helvetica-Oblique', fontSize=7.5,
        textColor=colors.HexColor('#555555'), spaceAfter=4,
    )

    flag_widths = [0.5*inch, 0.7*inch, 0.7*inch, 0.6*inch,
                   1.2*inch, 0.5*inch, 3.3*inch]
    mp_widths   = [1.8*inch, 0.8*inch, 1.5*inch, 1.2*inch, 1.2*inch]
    eb_widths   = [1.8*inch, 0.8*inch, 1.7*inch, 1.5*inch]
    audit_widths = [1.2*inch, 1.5*inch, 1.5*inch, 3.3*inch]
    qr_widths   = [0.8*inch, 0.75*inch, 0.75*inch, 0.7*inch,
                   0.7*inch, 0.7*inch, 1.4*inch, 0.5*inch, 0.7*inch]

    def section_elems(meta, ws_source=None, col_widths=None, include_table=True):
        rs, re_ = meta['row_start'], meta['row_end']
        cs, ce  = meta['col_start'], meta['col_end']
        count   = meta['count']
        elems = [
            Paragraph(str(meta['title'] or ""), style_section),
            Paragraph(str(meta['text']  or ""), style_body),
        ]
        if include_table:
            has_table = not (rs == re_ and count == 0)
            if has_table and all(v is not None for v in [rs, re_, cs, ce]):
                src  = ws_source or ws_email
                data = get_table_data(src, rs, re_, cs, ce)
                tbl  = make_table(data, col_widths=col_widths)
                if tbl:
                    elems.append(tbl)
        elems.append(Spacer(1, 4))
        return elems

    title_text = "PVD AA Cabin Report" + (" (Abridged)" if abridged else "")
    story = [
        Paragraph(title_text, style_title),
        Paragraph(date_line,  style_small),
        Spacer(1, 6),
    ]

    def add(meta, ws_source=None, col_widths=None, is_flags=False):
        show_table = (not abridged) or is_flags
        story.append(KeepTogether(
            section_elems(meta, ws_source=ws_source,
                          col_widths=col_widths, include_table=show_table)
        ))

    add(sections[2])
    add(lms_section)
    add(sections[4])
    add(sections[5])
    add(sections[6])
    add(sections[7])
    add(sections[8],  col_widths=flag_widths, is_flags=True)
    add(sections[9])
    add(sections[14], col_widths=mp_widths)
    add(sections[15], col_widths=eb_widths)
    add(sections[16], col_widths=audit_widths)
    add(sections[12], ws_source=ws_jobs, col_widths=qr_widths)

    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        leftMargin=0.4*inch, rightMargin=0.4*inch,
        topMargin=0.4*inch,  bottomMargin=0.4*inch,
    )
    doc.build(story)
    return buf.getvalue()


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    # 1. Authenticate
    print("Authenticating with Microsoft Graph...")
    token = get_access_token()

    # 2. Download PVD Excel from SharePoint
    excel_bytes = download_excel(token)
    wb       = openpyxl.load_workbook(excel_bytes, data_only=True)
    ws_email = wb['Email Generation']
    ws_jobs  = wb['Sheet1']          # PVD uses Sheet1 as the Jobs Email source

    # 3. Read submission ID and dates
    submission_id = str(ws_email['B1'].value or "report")
    if submission_id.startswith('_'):
        submission_id = submission_id[1:]

    start_date_raw = ws_email.cell(row=10, column=2).value
    end_date_raw   = ws_email.cell(row=11, column=2).value
    start_date_str = fmt_value(start_date_raw) if start_date_raw else ""
    end_date_str   = fmt_value(end_date_raw)   if end_date_raw   else ""

    date_line = start_date_str
    if end_date_str and end_date_str != start_date_str:
        date_line += f" \u2013 {end_date_str}"

    start_short = fmt_date_short(start_date_raw) if start_date_raw else ""
    end_short   = fmt_date_short(end_date_raw)   if end_date_raw   else ""
    date_short  = (f"{start_short}-{end_short}"
                   if end_short and end_short != start_short
                   else start_short)

    # 4. Read section metadata (rows 2–15 of Email Generation)
    sections = {}
    for sheet_row in range(2, 17):
        row_data = list(ws_email.iter_rows(
            min_row=sheet_row, max_row=sheet_row,
            min_col=1, max_col=7, values_only=True,
        ))[0]
        title, text, count, d, e, f, g = row_data
        sections[sheet_row] = {
            'title':     title,
            'text':      text,
            'count':     count or 0,
            'col_start': d,
            'row_start': e,
            'row_end':   f,
            'col_end':   g,
        }

    lms_section = (
        sections[13] if sections[13]['count'] < sections[3]['count']
        else sections[3]
    )

    # 5. Generate PDFs
    full_filename     = f"PVD Report Document_{submission_id}.pdf"
    abridged_filename = f"PVD Report Document Abridged_{submission_id}.pdf"

    print("Generating full report PDF...")
    full_bytes = build_pdf_bytes(
        ws_email, ws_jobs, date_line, sections, lms_section, abridged=False)

    print("Generating abridged report PDF...")
    abridged_bytes = build_pdf_bytes(
        ws_email, ws_jobs, date_line, sections, lms_section, abridged=True)

    # 6. Upload PDFs to SharePoint
    upload_pdf(token, full_filename,     full_bytes)
    upload_pdf(token, abridged_filename, abridged_bytes)

    # 7. Download screenshot for email body
    print("Downloading station screenshot...")
    screenshot_bytes = download_screenshot(token)

    # 8. Send email with both PDFs attached
    if SEND_EMAIL:
        subject = f"PVD Recap: {date_short}"
        body = (
            f"Hello team,\n\n"
            f"View the attached abridged and full reports for PVD on {date_short}. "
            f"Please reach out to Maren with any questions about the reported "
            f"operations, and reach out to Sam with questions about the report "
            f"creation or formatting.\n\n"
            f"Thanks,\n"
            f"Foxtrot Automation Services"
        )
        send_email(
            token=token,
            subject=subject,
            body_text=body,
            recipients=EMAIL_RECIPIENTS,
            from_address=EMAIL_FROM,
            attachments=[
                (abridged_filename, abridged_bytes),
                (full_filename,     full_bytes),
            ],
            screenshot_bytes=screenshot_bytes,
        )
    else:
        print("Email sending is disabled (SEND_EMAIL = False).")

    print("All done!")


if __name__ == "__main__":
    main()
